from __future__ import annotations

import json
import os
import re
import shutil
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import cv2
import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pdf2image import convert_from_bytes, convert_from_path

try:
    from paddleocr import PPStructure
    OCR_IMPORT_ERROR = None
except ImportError as exc:
    PPStructure = None
    OCR_IMPORT_ERROR = exc


CACHE_SUBDIR = ".pdf_cache"
RECORDS_SUBDIR = "records"
SAMPLES_SUBDIR = "samples"
PROGRESS_FILE = "progress.json"
OUTPUT_FILENAME = "Extracted_Data.xlsx"
MAX_SAMPLE_FILES = 6


st.set_page_config(
    page_title="Handwritten PDF to Excel",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    /* ── Base & background ── */
    .stApp {
        background:
            radial-gradient(ellipse 80% 50% at 10% 0%, rgba(20, 184, 166, 0.12) 0%, transparent 60%),
            radial-gradient(ellipse 60% 40% at 90% 5%, rgba(99, 102, 241, 0.10) 0%, transparent 55%),
            linear-gradient(180deg, #0D1117 0%, #0A0F18 100%);
    }
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2.5rem;
        max-width: 1280px;
    }

    /* ── Hero & info cards ── */
    .hero-card {
        background: linear-gradient(135deg, rgba(20,184,166,0.10) 0%, rgba(99,102,241,0.08) 100%);
        border: 1px solid rgba(20, 184, 166, 0.25);
        border-radius: 20px;
        padding: 1.4rem 1.6rem;
        box-shadow: 0 0 0 1px rgba(255,255,255,0.04), 0 20px 40px rgba(0,0,0,0.45);
        backdrop-filter: blur(12px);
        margin-bottom: 0.5rem;
    }
    .info-card {
        background: rgba(22, 27, 34, 0.85);
        border: 1px solid rgba(48, 54, 61, 0.9);
        border-radius: 16px;
        padding: 1rem 1.1rem;
        box-shadow: 0 8px 24px rgba(0,0,0,0.35);
        backdrop-filter: blur(8px);
        font-size: 0.9rem;
        color: #94A3B8;
        line-height: 1.65;
    }
    .info-card b {
        color: #E2E8F0;
        display: block;
        margin-bottom: 0.35rem;
        font-size: 0.95rem;
    }

    /* ── Hero text ── */
    .hero-title {
        font-size: 2.3rem;
        font-weight: 800;
        line-height: 1.1;
        letter-spacing: -0.02em;
        background: linear-gradient(135deg, #14B8A6 0%, #818CF8 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 0.4rem;
    }
    .hero-subtitle {
        color: #94A3B8;
        font-size: 1rem;
        margin-bottom: 0;
        line-height: 1.6;
    }

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] {
        background: #0D1117 !important;
        border-right: 1px solid rgba(48, 54, 61, 0.8) !important;
    }
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] label {
        color: #94A3B8 !important;
    }

    /* ── Buttons ── */
    .stButton > button {
        border-radius: 12px;
        height: 3rem;
        font-weight: 600;
        font-size: 0.95rem;
        letter-spacing: 0.01em;
        border: none;
        transition: all 0.2s ease;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #0D9488 0%, #0EA5E9 100%) !important;
        box-shadow: 0 4px 20px rgba(13, 148, 136, 0.4);
    }
    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 28px rgba(13, 148, 136, 0.6);
        transform: translateY(-1px);
    }
    .stButton > button:not([kind="primary"]) {
        background: rgba(30, 36, 46, 0.9) !important;
        border: 1px solid rgba(48, 54, 61, 0.9) !important;
        color: #94A3B8 !important;
    }
    .stButton > button:not([kind="primary"]):hover {
        border-color: rgba(20, 184, 166, 0.5) !important;
        color: #14B8A6 !important;
    }

    /* ── Metrics ── */
    div[data-testid="stMetric"] {
        background: rgba(22, 27, 34, 0.9);
        border: 1px solid rgba(48, 54, 61, 0.9);
        padding: 0.85rem 1rem;
        border-radius: 14px;
        box-shadow: 0 4px 16px rgba(0,0,0,0.3);
    }
    div[data-testid="stMetricLabel"] {
        color: #64748B !important;
        font-size: 0.8rem !important;
        font-weight: 500 !important;
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }
    div[data-testid="stMetricValue"] {
        font-size: 1.85rem !important;
        font-weight: 700 !important;
        color: #E2E8F0 !important;
    }

    /* ── Progress bar ── */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #0D9488, #0EA5E9) !important;
        border-radius: 99px;
    }
    .stProgress > div > div {
        background: rgba(30, 36, 46, 0.8) !important;
        border-radius: 99px;
    }

    /* ── File uploader ── */
    [data-testid="stFileUploader"] {
        background: rgba(22, 27, 34, 0.7);
        border: 1.5px dashed rgba(20, 184, 166, 0.35);
        border-radius: 16px;
        padding: 0.5rem;
        transition: border-color 0.2s;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: rgba(20, 184, 166, 0.65);
    }

    /* ── Expanders ── */
    details {
        background: rgba(22, 27, 34, 0.7) !important;
        border: 1px solid rgba(48, 54, 61, 0.8) !important;
        border-radius: 12px !important;
    }

    /* ── Divider ── */
    hr {
        border-color: rgba(48, 54, 61, 0.8) !important;
    }

    /* ── Alerts / banners ── */
    div[data-testid="stAlert"] {
        border-radius: 12px !important;
        border-left-width: 3px !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def cache_paths(output_folder: str) -> dict[str, Path]:
    base = Path(output_folder) / CACHE_SUBDIR
    records = base / RECORDS_SUBDIR
    samples = base / SAMPLES_SUBDIR
    base.mkdir(parents=True, exist_ok=True)
    records.mkdir(parents=True, exist_ok=True)
    samples.mkdir(parents=True, exist_ok=True)
    return {
        "base": base,
        "records": records,
        "samples": samples,
        "progress": base / PROGRESS_FILE,
    }


def default_input_path() -> str:
    if Path("/data/input").exists():
        return "/data/input"
    return str((Path.cwd() / "input_pdfs").resolve())


def default_output_path() -> str:
    if Path("/data/output").exists():
        return "/data/output"
    fallback = Path.cwd() / "output_excel"
    fallback.mkdir(parents=True, exist_ok=True)
    return str(fallback.resolve())


def file_token(filename: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", Path(filename).stem).strip("_") or "file"


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).replace("\n", " ").strip()
    if text.lower() == "nan":
        return ""
    return " ".join(text.split())


def count_meaningful_values(record: dict[str, str]) -> int:
    return sum(1 for value in record.values() if safe_text(value))


def preprocess_variants(image_bgr: np.ndarray) -> list[np.ndarray]:
    variants = [image_bgr]

    gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.2, tileGridSize=(8, 8)).apply(gray)
    denoised = cv2.fastNlMeansDenoising(clahe, None, h=10, templateWindowSize=7, searchWindowSize=21)
    thresh = cv2.adaptiveThreshold(
        denoised,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        11,
    )
    cleaned = cv2.cvtColor(thresh, cv2.COLOR_GRAY2BGR)

    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]], dtype=np.float32)
    sharpened = cv2.filter2D(cleaned, -1, kernel)

    variants.append(cv2.cvtColor(denoised, cv2.COLOR_GRAY2BGR))
    variants.append(cleaned)
    variants.append(sharpened)
    return variants


def dedupe_columns(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for idx, col in enumerate(columns, start=1):
        base = safe_text(col) or f"Extra_{idx}"
        counts[base] = counts.get(base, 0) + 1
        if counts[base] > 1:
            result.append(f"{base}_{counts[base]}")
        else:
            result.append(base)
    return result


def normalize_table(raw: pd.DataFrame) -> pd.DataFrame:
    table = raw.copy()
    table = table.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if table.empty:
        return pd.DataFrame()

    table = table.map(safe_text)
    table = table.replace("", np.nan).dropna(axis=0, how="all").dropna(axis=1, how="all")
    if table.empty:
        return pd.DataFrame()

    generic_headers = isinstance(table.columns, pd.RangeIndex) or all(
        safe_text(col).lower().startswith("unnamed") or not safe_text(col)
        for col in table.columns
    )

    if generic_headers and len(table.index) >= 2:
        first_row = [safe_text(v) for v in table.iloc[0].tolist()]
        meaningful = [v for v in first_row if v]
        if meaningful:
            table.columns = dedupe_columns(first_row)
            table = table.iloc[1:].reset_index(drop=True)
    else:
        table.columns = dedupe_columns([safe_text(col) for col in table.columns])

    table = table.map(safe_text)
    table = table.replace("", np.nan).dropna(axis=0, how="all").fillna("")
    return table.reset_index(drop=True)


def table_to_record(table: pd.DataFrame) -> dict[str, str]:
    if table.empty:
        return {}

    if len(table.columns) >= 2:
        first_col = table.iloc[:, 0].map(safe_text).tolist()
        second_col = table.iloc[:, 1:].apply(
            lambda row: " | ".join([safe_text(v) for v in row if safe_text(v)]),
            axis=1,
        ).tolist()
        label_count = sum(1 for item in first_col if item)
        if label_count >= max(2, len(table.index) - 1):
            record: dict[str, str] = {}
            extra_index = 1
            for label, value in zip(first_col, second_col, strict=False):
                clean_label = safe_text(label)
                clean_value = safe_text(value)
                if not clean_label and clean_value:
                    clean_label = f"Extra_{extra_index}"
                    extra_index += 1
                if clean_label:
                    record[clean_label] = clean_value
            if record:
                return record

    record = {}
    for column in table.columns:
        values = [safe_text(v) for v in table[column].tolist() if safe_text(v)]
        if values:
            record[column] = " | ".join(values)
    return record


def parse_regions_to_record(regions: list[dict[str, Any]]) -> tuple[dict[str, str], list[pd.DataFrame], list[str]]:
    merged_record: dict[str, str] = {}
    sample_tables: list[pd.DataFrame] = []
    floating_text: list[str] = []
    extra_index = 1

    for region in regions:
        region_type = region.get("type", "")

        if region_type == "table":
            html = region.get("res", {}).get("html", "")
            if not html:
                continue
            try:
                parsed = pd.read_html(StringIO(html))[0]
            except Exception:
                continue
            normalized = normalize_table(parsed)
            if normalized.empty:
                continue
            sample_tables.append(normalized.head(5))
            for key, value in table_to_record(normalized).items():
                clean_key = safe_text(key) or f"Extra_{extra_index}"
                if clean_key.startswith("Extra_"):
                    extra_index += 1
                clean_value = safe_text(value)
                if not clean_value:
                    continue
                if clean_key in merged_record and merged_record[clean_key] != clean_value:
                    merged_record[clean_key] = f"{merged_record[clean_key]} | {clean_value}"
                else:
                    merged_record[clean_key] = clean_value

        elif region_type in {"text", "title"}:
            text = " ".join(
                safe_text(line.get("text"))
                for line in region.get("res", [])
                if safe_text(line.get("text"))
            ).strip()
            if text:
                floating_text.append(text)

    for index, text in enumerate(floating_text, start=1):
        merged_record[f"Unstructured_{index}"] = text

    return merged_record, sample_tables, floating_text


@st.cache_resource(show_spinner="Loading OCR engine. The first run can take a minute.")
def load_engine():
    if OCR_IMPORT_ERROR is not None or PPStructure is None:
        raise RuntimeError(
            "PPStructure is not available in the installed paddleocr package. "
            "Rebuild the container after updating requirements."
        ) from OCR_IMPORT_ERROR

    return PPStructure(
        show_log=False,
        image_orientation=True,
        table=True,
        layout=True,
        ocr=True,
        use_gpu=False,
    )


def load_progress(progress_path: Path) -> dict[str, Any]:
    if progress_path.exists():
        return json.loads(progress_path.read_text(encoding="utf-8"))
    return {"files": {}, "excel_path": ""}


def save_progress(progress_path: Path, progress: dict[str, Any]) -> None:
    progress_path.write_text(json.dumps(progress, indent=2), encoding="utf-8")


def record_path(records_dir: Path, filename: str) -> Path:
    return records_dir / f"{file_token(filename)}.json"


def sample_path(samples_dir: Path, filename: str) -> Path:
    return samples_dir / f"{file_token(filename)}.png"


def save_record(records_dir: Path, filename: str, record: dict[str, str]) -> None:
    payload = {"file_name": filename, "record": record}
    record_path(records_dir, filename).write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_record(records_dir: Path, filename: str) -> dict[str, str] | None:
    path = record_path(records_dir, filename)
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload.get("record", {})


def save_sample_image(image, target: Path) -> None:
    sample = image.copy()
    sample.thumbnail((900, 900))
    sample.save(target, format="PNG")


def collect_saved_results(records_dir: Path, all_files: list[str]) -> list[tuple[str, dict[str, str]]]:
    results: list[tuple[str, dict[str, str]]] = []
    for filename in all_files:
        record = load_record(records_dir, filename)
        if record:
            results.append((filename, record))
    return results


def write_excel(results: list[tuple[str, dict[str, str]]], output_path: str) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Extracted Data"

    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    value_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 1
    max_columns = 1

    for filename, record in results:
        ordered_record = {"Source_File": filename, **record}
        headers = list(ordered_record.keys())
        values = [ordered_record[key] for key in headers]
        max_columns = max(max_columns, len(headers))

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        for col_idx, value in enumerate(values, start=1):
            cell = worksheet.cell(row=current_row + 1, column=col_idx, value=value)
            cell.alignment = value_align
            cell.border = border

        worksheet.row_dimensions[current_row].height = 24
        worksheet.row_dimensions[current_row + 1].height = 38
        current_row += 3

    for col_idx in range(1, max_columns + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        width = 14
        for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    width = min(max(width, len(str(cell.value)) + 2), 40)
        worksheet.column_dimensions[letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    excel_bytes = buffer.getvalue()
    Path(output_path).write_bytes(excel_bytes)
    return excel_bytes


def extract_from_pdf_bytes(engine, pdf_bytes: bytes, sample_target: Path) -> tuple[dict[str, str], list[pd.DataFrame]]:
    pages = convert_from_bytes(pdf_bytes, dpi=220, first_page=1, last_page=1)
    if not pages:
        return {}, []

    save_sample_image(pages[0], sample_target)
    image = cv2.cvtColor(np.array(pages[0]), cv2.COLOR_RGB2BGR)
    best_record: dict[str, str] = {}
    best_tables: list[pd.DataFrame] = []

    for variant in preprocess_variants(image):
        regions = engine(variant)
        record, sample_tables, _ = parse_regions_to_record(regions)
        if count_meaningful_values(record) > count_meaningful_values(best_record):
            best_record = record
            best_tables = sample_tables

    return best_record, best_tables


def extract_pdf_record(engine, pdf_path: str, samples_dir: Path) -> tuple[dict[str, str], list[pd.DataFrame]]:
    pages = convert_from_path(pdf_path, dpi=220, first_page=1, last_page=1)
    if not pages:
        return {}, []

    save_sample_image(pages[0], sample_path(samples_dir, Path(pdf_path).name))
    image = cv2.cvtColor(np.array(pages[0]), cv2.COLOR_RGB2BGR)
    best_record: dict[str, str] = {}
    best_tables: list[pd.DataFrame] = []

    for variant in preprocess_variants(image):
        regions = engine(variant)
        record, sample_tables, _ = parse_regions_to_record(regions)
        if count_meaningful_values(record) > count_meaningful_values(best_record):
            best_record = record
            best_tables = sample_tables

    return best_record, best_tables


def summarize_progress(progress: dict[str, Any], all_files: list[str]) -> tuple[int, int, int]:
    files = progress.get("files", {})
    done = sum(1 for item in all_files if files.get(item, {}).get("status") == "done")
    empty = sum(1 for item in all_files if files.get(item, {}).get("status") == "empty")
    errors = sum(1 for item in all_files if files.get(item, {}).get("status") == "error")
    return done, empty, errors


def error_rows(progress: dict[str, Any], all_files: list[str]) -> pd.DataFrame:
    files = progress.get("files", {})
    rows = []
    for filename in all_files:
        item = files.get(filename, {})
        if item.get("status") == "error":
            rows.append({"File": filename, "Error": item.get("message", "")})
    return pd.DataFrame(rows)


def process_uploaded_files(
    uploaded_files: list[Any],
    output_folder: str,
) -> tuple[list[tuple[str, dict[str, str]]], pd.DataFrame, bytes | None]:
    engine = load_engine()
    paths = cache_paths(output_folder)
    uploaded_files = sorted(uploaded_files, key=lambda item: item.name)

    if not uploaded_files:
        st.warning("Upload at least one PDF file.")
        return [], pd.DataFrame(), None

    c1, c2, c3 = st.columns(3)
    c1.metric("Uploaded PDFs", len(uploaded_files))
    c2.metric("Processing Mode", "Upload")
    c3.metric("Resume", "Off")

    st.info(
        "Upload mode is best for hosted Streamlit deployment. Files are processed in this run and the workbook is ready to download."
    )

    progress_bar = st.progress(0.0)
    live_status = st.empty()
    sample_area = st.container()
    results: list[tuple[str, dict[str, str]]] = []
    issues: list[dict[str, str]] = []

    for index, uploaded in enumerate(uploaded_files, start=1):
        live_status.info(f"Working on {uploaded.name} ({index}/{len(uploaded_files)})")
        try:
            record, sample_tables = extract_from_pdf_bytes(
                engine,
                uploaded.getvalue(),
                sample_path(paths["samples"], uploaded.name),
            )
            if record:
                results.append((uploaded.name, record))
                with sample_area:
                    with st.expander(f"Preview for {uploaded.name}", expanded=index <= 3):
                        if sample_tables:
                            st.dataframe(sample_tables[0], use_container_width=True, hide_index=True)
                        st.dataframe(
                            pd.DataFrame([{"Source_File": uploaded.name, **record}]),
                            use_container_width=True,
                            hide_index=True,
                        )
            else:
                issues.append({"File": uploaded.name, "Error": "No structured data extracted"})
        except Exception as exc:
            issues.append({"File": uploaded.name, "Error": str(exc)})
            st.warning(f"Skipped {uploaded.name}: {exc}")

        progress_bar.progress(index / len(uploaded_files))

    excel_bytes = None
    if results:
        excel_bytes = write_excel(results, str(Path(output_folder) / OUTPUT_FILENAME))

    live_status.success(
        f"Finished. Done: {len(results)}, errors: {len(issues)}. Download the workbook below."
    )
    return results, pd.DataFrame(issues), excel_bytes


def show_saved_samples(output_folder: str, results: list[tuple[str, dict[str, str]]]) -> None:
    if not results:
        return

    paths = cache_paths(output_folder)
    st.subheader("Samples")
    sample_columns = st.columns(2)

    for idx, (filename, record) in enumerate(results[:MAX_SAMPLE_FILES]):
        image_path = sample_path(paths["samples"], filename)
        with sample_columns[idx % 2]:
            with st.container():
                st.caption(filename)
                if image_path.exists():
                    st.image(str(image_path), use_container_width=True)
                preview_df = pd.DataFrame([{"Source_File": filename, **record}]).T.reset_index()
                preview_df.columns = ["Field", "Value"]
                st.dataframe(preview_df.head(12), use_container_width=True, hide_index=True)


def process_files(
    input_folder: str,
    output_folder: str,
    limit: int,
    resume: bool,
    include_errors_on_resume: bool,
) -> tuple[list[tuple[str, dict[str, str]]], dict[str, Any], bytes | None]:
    engine = load_engine()
    paths = cache_paths(output_folder)
    excel_path = str(Path(output_folder) / OUTPUT_FILENAME)

    all_files = sorted(
        file_name
        for file_name in os.listdir(input_folder)
        if file_name.lower().endswith(".pdf")
    )
    if limit > 0:
        all_files = all_files[:limit]

    if not all_files:
        st.warning("No PDF files were found in the selected input folder.")
        return [], {"files": {}, "excel_path": excel_path}, None

    if not resume:
        if paths["base"].exists():
            shutil.rmtree(paths["base"])
        paths = cache_paths(output_folder)
        if Path(excel_path).exists():
            Path(excel_path).unlink()
        progress = {"files": {}, "excel_path": excel_path}
    else:
        progress = load_progress(paths["progress"])
        progress["excel_path"] = excel_path

    files_progress = progress.setdefault("files", {})
    done_count, empty_count, error_count = summarize_progress(progress, all_files)

    todo: list[str] = []
    for filename in all_files:
        status = files_progress.get(filename, {}).get("status")
        if status == "done":
            continue
        if status == "empty":
            continue
        if status == "error" and not include_errors_on_resume:
            continue
        todo.append(filename)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total PDFs", len(all_files))
    c2.metric("Done", done_count)
    c3.metric("Empty", empty_count)
    c4.metric("Errors", error_count)

    if todo:
        st.info(
            f"Processing {len(todo)} file(s). The Excel file is auto-saved to {excel_path} after each success."
        )
    else:
        st.success("Nothing new to process. Loading the saved records from cache.")

    progress_bar = st.progress((done_count + empty_count + error_count) / len(all_files))
    live_status = st.empty()
    sample_area = st.container()
    excel_bytes: bytes | None = None

    for idx, filename in enumerate(todo, start=1):
        current_number = len(all_files) - len(todo) + idx
        live_status.info(f"Working on {filename} ({current_number}/{len(all_files)})")
        pdf_path = str(Path(input_folder) / filename)

        try:
            record, sample_tables = extract_pdf_record(engine, pdf_path, paths["samples"])
            if record:
                save_record(paths["records"], filename, record)
                files_progress[filename] = {"status": "done", "columns": list(record.keys())}
                results = collect_saved_results(paths["records"], all_files)
                excel_bytes = write_excel(results, excel_path)

                with sample_area:
                    with st.expander(f"Sample for {filename}", expanded=idx <= 3):
                        if sample_tables:
                            st.dataframe(sample_tables[0], use_container_width=True, hide_index=True)
                        preview_row = pd.DataFrame([{"Source_File": filename, **record}])
                        st.dataframe(preview_row, use_container_width=True, hide_index=True)
            else:
                files_progress[filename] = {"status": "empty"}

        except Exception as exc:
            files_progress[filename] = {"status": "error", "message": str(exc)}
            st.warning(f"Skipped {filename}: {exc}")

        save_progress(paths["progress"], progress)
        done_count, empty_count, error_count = summarize_progress(progress, all_files)
        progress_bar.progress((done_count + empty_count + error_count) / len(all_files))

    final_results = collect_saved_results(paths["records"], all_files)
    final_done, final_empty, final_errors = summarize_progress(progress, all_files)
    if excel_bytes is None and Path(excel_path).exists():
        excel_bytes = Path(excel_path).read_bytes()
    live_status.success(
        f"Finished. Done: {final_done}, empty: {final_empty}, errors: {final_errors}. "
        f"Latest Excel saved to {excel_path}."
    )
    return final_results, progress, excel_bytes


with st.sidebar:
    st.title("Run Settings")
    run_mode = st.radio(
        "Input mode",
        options=["Upload PDFs", "Read Folder"],
        help="Use uploads for hosted Streamlit. Use folder mode for Docker or private server installs.",
    )
    output_path = st.text_input(
        "Output folder path",
        value=default_output_path(),
        help="Folder where the Excel file, progress cache, and samples will be saved.",
    )

    if run_mode == "Read Folder":
        input_path = st.text_input(
            "PDF folder path",
            value=default_input_path(),
            help="Folder that contains your one-page PDF files.",
        )
        limit = st.number_input(
            "File limit",
            min_value=0,
            value=0,
            help="Use 0 to process all files. Any positive number is useful for quick testing.",
        )
        resume_mode = st.toggle(
            "Resume previous run",
            value=True,
            help="Skips files already marked done or empty in the cache.",
        )
        retry_errors = st.toggle(
            "Retry failed files",
            value=True,
            help="When enabled, files marked as error are retried on the next run.",
        )
    else:
        input_path = ""
        limit = 0
        resume_mode = False
        retry_errors = True

    st.markdown(
        """
        <div class="info-card">
            <b>How it works</b><br>
            Upload mode is best for hosted Streamlit apps.<br>
            Folder mode is best for Docker and private servers.<br><br>
            Each PDF becomes:<br>
            header row, data row, blank row.
        </div>
        """,
        unsafe_allow_html=True,
    )


st.markdown(
    """
    <div class="hero-card">
        <div class="hero-title">PDF to Excel Extractor</div>
        <p class="hero-subtitle">
            Streamlit-friendly OCR for scanned technical tables and one-page PDF images.
            Use uploads for hosted deployment or folders for Docker and server runs.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

mode_col1, mode_col2 = st.columns([2, 1])
with mode_col1:
    st.markdown(
        """
        <div class="info-card">
            <b>Smart workflow</b><br>
            Hosted Streamlit deployment: upload PDFs directly in the browser.<br>
            Private server or Docker: point the app to an input folder and keep resume enabled.
        </div>
        """,
        unsafe_allow_html=True,
    )
with mode_col2:
    st.metric("Current Mode", run_mode)

uploaded_files: list[Any] = []
if run_mode == "Upload PDFs":
    uploaded_files = st.file_uploader(
        "Upload one or more PDF files",
        type=["pdf"],
        accept_multiple_files=True,
        help="Best option for Streamlit deployment.",
    )

start_col, clear_col = st.columns([4, 1])
start_clicked = start_col.button("Start Extraction", type="primary", use_container_width=True)
clear_clicked = clear_col.button("Clear Cache", use_container_width=True)

if clear_clicked:
    cache_dir = Path(output_path) / CACHE_SUBDIR
    if cache_dir.exists():
        shutil.rmtree(cache_dir)
        st.success("Cache cleared. The next run will start fresh.")
    else:
        st.info("No cache folder was found.")

if OCR_IMPORT_ERROR is not None:
    st.error(
        "The installed paddleocr package does not expose PPStructure, so the OCR engine cannot start. "
        "Rebuild the Docker image after updating the pinned requirements."
    )
    st.code(str(OCR_IMPORT_ERROR))
    st.stop()

if start_clicked:
    Path(output_path).mkdir(parents=True, exist_ok=True)

    if run_mode == "Read Folder":
        path_errors: list[str] = []
        if not os.path.isdir(input_path):
            path_errors.append(f"Input folder not found: {input_path}")
        if not os.path.isdir(output_path):
            path_errors.append(f"Output folder not found: {output_path}")

        if path_errors:
            for message in path_errors:
                st.error(message)
            st.stop()

        all_files = sorted(
            file_name for file_name in os.listdir(input_path) if file_name.lower().endswith(".pdf")
        )
        if int(limit) > 0:
            all_files = all_files[: int(limit)]

        try:
            results, progress, excel_bytes = process_files(
                input_folder=input_path,
                output_folder=output_path,
                limit=int(limit),
                resume=resume_mode,
                include_errors_on_resume=retry_errors,
            )
        except Exception as exc:
            st.exception(exc)
            st.stop()

        if not results:
            st.warning("The run completed, but no structured records were extracted.")
            st.stop()

        excel_file = Path(output_path) / OUTPUT_FILENAME
        done_count, empty_count, error_count = summarize_progress(progress, all_files)

        st.divider()
        st.subheader("Result")

        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Saved records", len(results))
        col_b.metric("Empty files", empty_count)
        col_c.metric("Errors", error_count)

        st.success(f"Excel saved automatically to {excel_file}")

        if excel_bytes is not None:
            st.download_button(
                label="Download Excel",
                data=excel_bytes,
                file_name=OUTPUT_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        issue_df = error_rows(progress, all_files)
        if not issue_df.empty:
            st.subheader("Files Needing Attention")
            st.dataframe(issue_df, use_container_width=True, hide_index=True)

        show_saved_samples(output_path, results)

    else:
        if not uploaded_files:
            st.error("Upload at least one PDF before starting.")
            st.stop()

        try:
            results, issue_df, excel_bytes = process_uploaded_files(uploaded_files, output_path)
        except Exception as exc:
            st.exception(exc)
            st.stop()

        if not results:
            st.warning("The run completed, but no structured records were extracted.")
            if not issue_df.empty:
                st.dataframe(issue_df, use_container_width=True, hide_index=True)
            st.stop()

        st.divider()
        st.subheader("Result")

        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Extracted records", len(results))
        col_b.metric("Uploaded files", len(uploaded_files))
        col_c.metric("Errors", len(issue_df))

        st.success(
            f"Excel is ready for download and also saved to {Path(output_path) / OUTPUT_FILENAME}"
        )

        if excel_bytes is not None:
            st.download_button(
                label="Download Excel",
                data=excel_bytes,
                file_name=OUTPUT_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        if not issue_df.empty:
            st.subheader("Files Needing Attention")
            st.dataframe(issue_df, use_container_width=True, hide_index=True)

        show_saved_samples(output_path, results)
